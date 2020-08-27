import logging
from os.path import split

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from box2adls.exceptions import XlsxFormatError
import box2adls.logging


logger = logging.getLogger(__name__)


def transform_daily(months, labels, source):
    """
    Transform daily Excel file.

    :param months: Names of tabs
    :param labels: New names of tabs
    :param source: path to file (xlsx)
    :return: None
    """
    logger.info(f'Renaming worksheet tabs for {split(source)[1]}')

    wb = load_workbook(filename=source)

    for name, label in zip(months, labels):
        if name in wb:
            ws = wb[name]
            ws.title = label
        else:
            raise XlsxFormatError(f"Invalid file format: {split(source)[1]}")

    wb.save(filename=source)


def transform_weekly(tab_name, tab_rename, source):
    """
    Transform weekly Excel file.

    :param tab_name: name of hidden tab
    :param tab_rename: new tab name
    :param source: path to file (xlsx)
    :return: None
    """
    logger.info(f'Enabling worksheet for {split(source)[1]}')

    wb = load_workbook(filename=source)

    if tab_name in wb:
        ws = wb[tab_name]
        ws.title = tab_rename
        ws.sheet_state = Worksheet.SHEETSTATE_VISIBLE
    else:
        raise XlsxFormatError(f"Invalid file format: {split(source)[1]}")

    wb.save(filename=source)
